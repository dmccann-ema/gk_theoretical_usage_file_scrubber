import logging
from sys import exit

import pandas as pd
from tkinter import Tk, filedialog
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from theoretical_use_file_scrubber import app
from . import utility
from .popup import start_loading

LOGGER = logging.getLogger()


def scrub_excel_file(excel_file_path):
    # Read the Excel file into a pandas DataFrame
    try:
        df = pd.read_excel(excel_file_path)
    except FileNotFoundError:
        print(f"Error: The file '{excel_file_path}' was not found.")
        exit(1)

    df.columns = df.iloc[3]
    
    df['Sold'] = pd.to_numeric(df['Sold'], errors='coerce', downcast='float')
    df['Use Each'] = pd.to_numeric(df['Use Each'], errors='coerce')
    df['Use Total'] = pd.to_numeric(df['Use Total'], errors='coerce')
    
    df = df[4:]
    
    first_col = df['Menu Item Name']
    unit_col = df['Unit']
    sold_col = df['Sold']
    use_col = df['Use Each']

    cleaned_df = df[
        first_col.notna() & 
        (first_col != '') & 
        (first_col != df.columns[0]) & 
        unit_col.notna() & 
        sold_col.notna() & 
        use_col.notna()
    ].dropna(axis='columns', how='all')

    file_path_parts = excel_file_path.split('.')

    excel_file_path_scrubbed = '.'.join(file_path_parts[:-1] + ['scrubbed', file_path_parts[-1]])

    with pd.ExcelWriter(excel_file_path_scrubbed, engine='openpyxl') as writer:
        cleaned_df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Grab the worksheet object
        ws = writer.sheets['Sheet1']

        # ---- Left‑align **only** the header row (row 1) ----
        for col_num in range(1, len(cleaned_df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.alignment = Alignment(horizontal='left')

        # (Optional) Auto‑size columns so you can see the effect
        for i, col in enumerate(cleaned_df.columns, start=1):
            col_letter = get_column_letter(i)
            max_len = max(
                len(str(s)) for s in cleaned_df[col]
            ) + 2                     # a little padding
            ws.column_dimensions[col_letter].width = max_len
        
    return excel_file_path_scrubbed


def select_file():
    return start_loading(
        lambda: filedialog.askopenfilename(
            title = "Select a file",
            filetypes = [("Excel files", "*.xls?"),]
        ),
        title="Working...",
        text="Please, select a file to scrub..."
    ).result


def scrub_file(excel_file_path, root=None):
    return start_loading(
        lambda: scrub_excel_file(
            excel_file_path=excel_file_path
        ), 
        title="Scrubbing file...", 
        text="Please wait, scrubbing...",
        root=root,
    ).result

def run(root=None):
    file_path = select_file()

    if not file_path:
        LOGGER.warning("No file selected. Operation canceled.")
        scrubbed_file_path = None

    else:
        LOGGER.debug(f"File selected: {file_path}")
        scrubbed_file_path = scrub_file(file_path, root)
        LOGGER.info(f"Excel file '{file_path}' has been scrubbed and saved to '{scrubbed_file_path}'.")
    
    
    if root and root.winfo_exists():
        root.grab_release()
        root.destroy()
    
    LOGGER.info("Closing scrubber.")
    exit(0)


def main(debug=False):
    if debug:
        utility.setDebugLogger(LOGGER)
        LOGGER.debug("Debugging enabled.")
    
    LOGGER.info(f"You are using {app.APPNAME}.")

    root = Tk()
    root.title(app.APPNAME)
    root.geometry("300x200")
    
    root.after(100, run, root)
    
    root.mainloop()
        
    LOGGER.info("Closing scrubber.")


if __name__ == "__main__":
    main()